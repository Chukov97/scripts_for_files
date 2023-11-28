from zipfile import ZipFile
from pypdf import PdfReader
from config import ZIP_FILE_DIR, LIST_FILE
from io import TextIOWrapper
from openpyxl import load_workbook
import os
import csv


def test_create_zip(create_zip):
    with ZipFile(ZIP_FILE_DIR, mode='a') as zf:
        num_files = len(zf.filelist)
        list_files = []
        for file in zf.infolist():
            name = os.path.basename(file.filename)
            list_files.append(name)
    assert num_files == 3
    assert LIST_FILE == list_files


def test_csv(create_zip):
    with ZipFile(ZIP_FILE_DIR) as zf:
        file_info = zf.getinfo('username.csv')
        file_size = file_info.file_size
        with zf.open('username.csv') as tmp_file:
            reader = list(csv.reader(TextIOWrapper(tmp_file, encoding='utf-8'), delimiter=';'))
        assert reader[1] == ['booker12', '9012', 'Rachel', 'Booker']
        assert file_size == 174


def test_pdf(create_zip):
    with ZipFile(ZIP_FILE_DIR) as zf:
        file_info = zf.getinfo('sample.pdf')
        file_size = file_info.file_size
        with zf.open('sample.pdf') as tmp_file:
            reader = PdfReader(tmp_file)
            page = reader.pages[0]
            text = page.extract_text()
            number_of_pages = len(reader.pages)
        assert number_of_pages == 2
        assert file_size == 3028
        assert 'just for use in the Virtual Mechanics tutorials' in text


def test_xlsx(create_zip):
    with ZipFile(ZIP_FILE_DIR) as zf:
        file_info = zf.getinfo('file_example_XLSX_50.xlsx')
        file_size = file_info.file_size
        with zf.open('file_example_XLSX_50.xlsx') as tmp_file:
            reader = load_workbook(tmp_file)
            sheet = reader.active
            print()
        assert sheet.cell(row=2, column=2).value == 'Dulce'
        assert file_size == 7360
